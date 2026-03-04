#!/usr/bin/env python3
"""
DICOM CSV Tag Processor

Processes CSV files containing DICOM tags, organizing them by series and instance numbers.
Supports flexible sorting by any DICOM tag value.
"""

import csv
import json
import os
import shutil
import sys
from typing import Dict, List, Optional, Any
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pydicom
    HAS_PYDICOM = True
except ImportError:
    HAS_PYDICOM = False
    print("Warning: pydicom not available. DICOM file parsing will be disabled.")


def get_ds(ds, group, elem) -> str:
    """Helper function to extract DICOM tag values safely."""
    try:
        v = str(ds[group, elem].value).upper()
        v = v.replace(',', ' ')   # remove all commas so it doesn't mess up the csv
        v = v.replace('\\', ' ')
        return str(v)
    except:
        return '-1'  # was 'Unk'


def get_kernel(ds) -> str:
    """
    Extract reconstruction kernel type from DICOM data structure.
    
    :param ds: pydicom data struct or JSON data dict
    :return: kernel classification: 'SMTH', 'STND', 'SHRP', or 'UNKN'
    """
    SOFT_KERNELS = ['BRAIN', 'SOFT_TISSUE', 'STANDARD', 'SOFT', 'SOFT TISSUE','A','B10F','B20F', 'B10', 'B20','FC08', 'FC12'] 
    MID_KERNELS = ['B30F', 'B35F', 'B45F', 'B40F', 'B30', 'B35', 'B40', 'FC17', 'FC18', 'FC24', 'B', 'Q34F', 'J30S', 'J30F']   
    HARD_KERNELS = ['LUNG', 'BONE', 'DETAIL', 'EDGE', 'SHARP', 'ULTRA', 'BONEPLUS', 'BONE PLUS', 'BONE+', 'E',\
                    'B50F','FC24','L','B60F','FC08','D','B70F','FC50','B80F','FC30','FC81']

    # Handle both pydicom dataset and JSON dict input
    if isinstance(ds, dict):
        # JSON data from CSV
        recon = ds.get('ConvolutionKernel', '') or ds.get('ReconstructionAlgorithm', '') or ds.get('FilterType', '')
        if isinstance(recon, str):
            recon = recon.upper()
        else:
            recon = ''
    else:
        # Try pydicom dataset format - first try tag 0x0018, 0x9314
        try:
            recon = ds[0x0018, 0x9314].value.upper()
        except:
            try:
                # Try tag 0x0018, 0x1210
                recon = ds[0x0018, 0x1210].value.upper()
            except:
                recon = ''

    # Check kernel categories
    if recon in SOFT_KERNELS:
        return 'SMTH'  # smooth
    elif recon in MID_KERNELS:
        return 'STND'  # standard
    elif recon in HARD_KERNELS:
        return 'SHRP'  # sharp

    # If no exact match, try numeric extraction
    if recon:
        try:
            # Extract digits from the reconstruction string
            value = ''.join([i for i in recon if i.isdigit()])
            if value:
                int_value = int(value)
                if int_value < 30:
                    return 'SMTH'
                elif int_value < 50:
                    return 'STND'
                else:
                    return 'SHRP'
        except:
            pass
    
    return 'UNKN'


def get_contrast(ds) -> str:
    """
    Determine contrast phase from DICOM data structure.
    
    :param ds: pydicom data struct or JSON data dict
    :return: 'Pre' or 'Pst' depending on if IV contrast has been given
    """
    # Assume no contrast, and change if any tag found saying contrast
    label = 'Pre'
    
    # Handle both pydicom dataset and JSON dict input
    if isinstance(ds, dict):
        # Check JSON data for contrast indicators
        contrast_agent = ds.get('ContrastBolusAgent', '')
        contrast_route = ds.get('ContrastBolusRoute', '')  
        contrast_volume = ds.get('ContrastBolusVolume', '')
        
        if (contrast_agent and str(contrast_agent).upper() not in ['', '-1', 'NONE']) or \
           (contrast_route and str(contrast_route).upper() not in ['', '-1', 'NONE']) or \
           (contrast_volume and str(contrast_volume).upper() not in ['', '-1', 'NONE']):
            label = 'Pst'
    else:
        # Try pydicom dataset format - check multiple contrast-related tags
        try:
            if ds[0x0018, 0x1044].value != '':  # Contrast/Bolus Agent
                label = 'Pst'
        except:
            pass
        try:
            if ds[0x0018, 0x0010].value != '':  # Contrast/Bolus Agent Sequence
                label = 'Pst'
        except:
            pass
        try:
            if ds[0x0018, 0x1040].value != '':  # Contrast/Bolus Route
                label = 'Pst'
        except:
            pass
        try:
            if ds[0x0018, 0x1048].value != '':  # Contrast/Bolus Volume
                label = 'Pst'
        except:
            pass
    
    return label


def getMajorAxisFromDirCos(x, y, z):
    """Get major axis from direction cosines."""
    axis = "UNK"
    if (x < 0):
        XOrient = "R"
    else:
        XOrient = "L"
    if (y < 0):
        YOrient = "A"
    else:
        YOrient = "P"
    if (z < 0):
        ZOrient = "F"
    else:
        ZOrient = "H"

    if x > 0:
        absX = x
    else:
        absX = -x
    if y > 0:
        absY = y
    else:
        absY = -y
    if z > 0:
        absZ = z
    else:
        absZ = -z

    if ((absX > 0.25) and (absX > absY) and (absX > absZ)):
        axis = XOrient
    elif ((absY > 0.25) and (absY > absX) and (absY > absZ)):
        axis = YOrient
    elif ((absZ > 0.25) and (absZ > absX) and (absZ > absY)):
        axis = ZOrient
    return axis


def getOrient(ds):
    """
    Get image orientation from DICOM data structure.
    
    :param ds: pydicom data struct or JSON data dict
    :return: returns the orientation of the slice: [ 'AXL' | 'COR' | 'SAG' | 'OBL' | 'UNK' ]
    """
    label = 'UNK'
    try:
        # Handle both pydicom dataset and JSON dict input
        if isinstance(ds, dict):
            # JSON data from CSV - parse ImageOrientationPatient
            image_orient = ds.get('ImageOrientationPatient', '')
            if isinstance(image_orient, str) and image_orient:
                # Parse string representation like "[1, 0, 0, 0, 1, 0]"
                import ast
                try:
                    if image_orient.startswith('[') and image_orient.endswith(']'):
                        ImageOrientDirCos = ast.literal_eval(image_orient)
                    else:
                        # Try splitting by comma or space
                        ImageOrientDirCos = [float(x.strip()) for x in image_orient.replace('[', '').replace(']', '').split(',')]
                except:
                    return 'UNK'
            elif isinstance(image_orient, (list, tuple)):
                ImageOrientDirCos = image_orient
            else:
                return 'UNK'
        else:
            # pydicom dataset format
            ImageOrientDirCos = ds[0x20, 0x0037].value
        
        # Ensure we have 6 values
        if len(ImageOrientDirCos) != 6:
            return 'UNK'
            
        rowAxis = getMajorAxisFromDirCos(ImageOrientDirCos[0], ImageOrientDirCos[1], ImageOrientDirCos[2])
        colAxis = getMajorAxisFromDirCos(ImageOrientDirCos[3], ImageOrientDirCos[4], ImageOrientDirCos[5])

        if (rowAxis != "" and colAxis != ""):
            if ((rowAxis == "R" or rowAxis == "L") and (colAxis == "A" or colAxis == "P")):
                label = "AXL"
            elif ((rowAxis == "R" or rowAxis == "L") and (colAxis == "H" or colAxis == "F")):
                label = "COR"
            elif ((rowAxis == "A" or rowAxis == "P") and (colAxis == "H" or colAxis == "F")):
                label = "SAG"
            else:
                label = "OBL"
        else:
            label = "OBL"
    except:
        label = "UNK"
    return label


def has_exclusion_keywords(series_description: str) -> bool:
    """
    Check if series description contains exclusion keywords that should be deprioritized.
    
    :param series_description: The series description string
    :return: True if contains exclusion keywords, False otherwise
    """
    EXCLUSION_KEYWORDS = ['SUB', 'SUBTRACTION', 'MIP', 'SCOUT', 'PROTOCOL', 'DOSE']
    
    if not series_description or not isinstance(series_description, str):
        return False
    
    desc_upper = series_description.upper()
    return any(keyword in desc_upper for keyword in EXCLUSION_KEYWORDS)


class DICOMImage:
    """Represents a single DICOM image with its associated tags."""
    
    REQUIRED_TAGS = [
        'SOPClassUID', 'ImageOrientationPatient', 'SliceThickness', 'ReconstructionAlgorithm',
        'ContrastBolusAgent', 'ContrastBolusRoute', 'ContrastBolusVolume', 'SeriesDescription',
        'ExamDescription', 'Modality', 'ConversionType', 'SeriesNumber', 'InstanceNumber',
        'BodyPartExamined', 'SC_Date', 'SliceLocation', 'Manufacturer', 'ManufacturerModelName',
        'ImagePositionPatient'
    ]
    
    def __init__(self, image_data: Dict[str, Any], json_data: Optional[Dict[str, Any]] = None, source_path: Optional[str] = None):
        """Initialize DICOM image with tag data."""
        self.tags = {}
        self.json_data = json_data or {}
        self.source_path = source_path  # Original DICOM file path (if available)
        
        for tag in self.REQUIRED_TAGS:
            self.tags[tag] = image_data.get(tag, None)
    
    def get_tag(self, tag_name: str) -> Any:
        """Get value for a specific DICOM tag."""
        return self.tags.get(tag_name)
    
    def get_instance_number(self) -> int:
        """Get instance number, defaulting to 0 if not present."""
        instance_num = self.tags.get('InstanceNumber')
        if instance_num is None:
            return 0
        try:
            return int(instance_num)
        except (ValueError, TypeError):
            return 0
    
    def get_series_number(self) -> int:
        """Get series number, defaulting to 0 if not present."""
        series_num = self.tags.get('SeriesNumber')
        if series_num is None:
            return 0
        try:
            return int(series_num)
        except (ValueError, TypeError):
            return 0
    
    def is_axial(self) -> bool:
        """Check if this image is from an axial series using ImageOrientationPatient."""
        # Create a combined dataset from both tags and json_data for the orientation function
        combined_data = {}
        
        # Add ImageOrientationPatient from tags
        if 'ImageOrientationPatient' in self.tags:
            combined_data['ImageOrientationPatient'] = self.tags['ImageOrientationPatient']
        
        # Add ImageOrientationPatient from JSON data (will override if present)
        if 'ImageOrientationPatient' in self.json_data:
            combined_data['ImageOrientationPatient'] = self.json_data['ImageOrientationPatient']
        
        # Use the improved orientation detection function
        orientation = getOrient(combined_data)
        return orientation == 'AXL'
    
    def get_kernel_hardness(self) -> str:
        """Get kernel hardness classification using improved kernel detection."""
        kernel_type = get_kernel(self.json_data)
        
        # Map the improved classifications to the display format
        if kernel_type == 'SHRP':
            return 'Hard'
        elif kernel_type == 'SMTH':
            return 'Soft'
        elif kernel_type == 'STND':
            return 'Standard'
        else:
            return 'Unknown'
    
    def get_spacing_between_slices(self) -> Optional[float]:
        """Get spacing between slices from JSON data."""
        spacing = self.json_data.get('SpacingBetweenSlices')
        if spacing:
            try:
                return float(spacing)
            except (ValueError, TypeError):
                pass
        return None
    
    def get_series_instance_uid(self) -> Optional[str]:
        """Get SeriesInstanceUID from JSON data."""
        uid = self.json_data.get('SeriesInstanceUID')
        if uid and isinstance(uid, str):
            return uid
        return None
    
    def get_orientation(self) -> str:
        """Get orientation classification: AXL, COR, SAG, OBL, or UNK."""
        combined_data = {}
        if 'ImageOrientationPatient' in self.tags:
            combined_data['ImageOrientationPatient'] = self.tags['ImageOrientationPatient']
        if 'ImageOrientationPatient' in self.json_data:
            combined_data['ImageOrientationPatient'] = self.json_data['ImageOrientationPatient']
        return getOrient(combined_data)
    
    def get_iv_contrast(self) -> str:
        """Get IV contrast status: 'YES' if post-contrast, 'NO' if pre-contrast."""
        combined_data = {}
        for key in ('ContrastBolusAgent', 'ContrastBolusRoute', 'ContrastBolusVolume'):
            if key in self.json_data:
                combined_data[key] = self.json_data[key]
            elif key in self.tags and self.tags[key]:
                combined_data[key] = self.tags[key]
        contrast = get_contrast(combined_data)
        return 'YES' if contrast == 'Pst' else 'NO'
    
    def __repr__(self):
        return f"DICOMImage(Series: {self.get_series_number()}, Instance: {self.get_instance_number()})"


class DICOMSeries:
    """Represents a DICOM series containing multiple images."""
    
    def __init__(self, series_number: int):
        """Initialize DICOM series."""
        self.series_number = series_number
        self.images: List[DICOMImage] = []
    
    def add_image(self, image: DICOMImage):
        """Add an image to this series."""
        self.images.append(image)
    
    def sort_by_instance_number(self):
        """Sort images by instance number."""
        self.images.sort(key=lambda img: img.get_instance_number())
    
    def sort_by_tag(self, tag_name: str, reverse: bool = False):
        """Sort images by any DICOM tag value."""
        def sort_key(img):
            value = img.get_tag(tag_name)
            if value is None:
                return ""
            return str(value)
        
        self.images.sort(key=sort_key, reverse=reverse)
    
    def __len__(self):
        return len(self.images)
    
    def is_axial(self) -> bool:
        """Check if this series contains axial images."""
        return any(img.is_axial() for img in self.images)
    
    def get_z_extent(self) -> Optional[float]:
        """Calculate z-extent (range) of the series in millimeters."""
        slice_locations = []
        for img in self.images:
            slice_loc = img.get_tag('SliceLocation')
            if slice_loc is not None:
                try:
                    slice_locations.append(float(slice_loc))
                except (ValueError, TypeError):
                    continue
        
        if len(slice_locations) >= 2:
            return max(slice_locations) - min(slice_locations)
        return None
    
    def get_kernel_hardness(self) -> str:
        """Get the predominant kernel hardness for this series."""
        hardness_counts = {'Hard': 0, 'Standard': 0, 'Soft': 0, 'Unknown': 0}
        for img in self.images:
            hardness = img.get_kernel_hardness()
            hardness_counts[hardness] += 1
        
        # Return the most common hardness
        return max(hardness_counts, key=hardness_counts.get)
    
    def get_average_slice_spacing(self) -> Optional[float]:
        """Get average slice spacing computed from sorted slice positions.
        
        Computes actual spacing by sorting slices by position and averaging
        the distances between adjacent slices. This is more accurate than
        the SpacingBetweenSlices DICOM header which can be incorrect.
        """
        # Method 1: Compute from SliceLocation (most reliable)
        slice_locations = []
        for img in self.images:
            slice_loc = img.get_tag('SliceLocation')
            if slice_loc is not None:
                try:
                    slice_locations.append(float(slice_loc))
                except (ValueError, TypeError):
                    continue
        
        if len(slice_locations) >= 2:
            slice_locations.sort()
            spacings = [slice_locations[i+1] - slice_locations[i] 
                        for i in range(len(slice_locations) - 1)]
            # Use absolute values in case of descending order
            spacings = [abs(s) for s in spacings if abs(s) > 0.001]
            if spacings:
                return sum(spacings) / len(spacings)
        
        # Method 2: Compute from ImagePositionPatient Z-coordinate
        z_positions = []
        for img in self.images:
            pos = img.json_data.get('ImagePositionPatient')
            if pos:
                try:
                    if isinstance(pos, (list, tuple)) and len(pos) >= 3:
                        z_positions.append(float(pos[2]))
                    elif isinstance(pos, str):
                        import ast
                        parsed = ast.literal_eval(pos)
                        if len(parsed) >= 3:
                            z_positions.append(float(parsed[2]))
                except (ValueError, TypeError, SyntaxError):
                    continue
        
        if len(z_positions) >= 2:
            z_positions.sort()
            spacings = [z_positions[i+1] - z_positions[i] 
                        for i in range(len(z_positions) - 1)]
            spacings = [abs(s) for s in spacings if abs(s) > 0.001]
            if spacings:
                return sum(spacings) / len(spacings)
        
        # Method 3: Fall back to SpacingBetweenSlices header
        spacings = []
        for img in self.images:
            spacing = img.get_spacing_between_slices()
            if spacing is not None:
                spacings.append(spacing)
        
        if spacings:
            return sum(spacings) / len(spacings)
        
        # Method 4: Fall back to SliceThickness
        thicknesses = []
        for img in self.images:
            thickness = img.get_tag('SliceThickness')
            if thickness is not None:
                try:
                    thicknesses.append(float(thickness))
                except (ValueError, TypeError):
                    continue
        
        if thicknesses:
            return sum(thicknesses) / len(thicknesses)
        
        return None
    
    def get_series_description(self) -> str:
        """Get series description from first image."""
        if self.images:
            desc = self.images[0].get_tag('SeriesDescription')
            if desc and isinstance(desc, str):
                return desc
        return f"Series {self.series_number}"
    
    def has_exclusion_keywords(self) -> bool:
        """Check if this series has exclusion keywords in its description."""
        series_desc = self.get_series_description()
        return has_exclusion_keywords(series_desc)
    
    def get_series_instance_uid(self) -> Optional[str]:
        """Get SeriesInstanceUID from the first image in this series."""
        if self.images:
            return self.images[0].get_series_instance_uid()
        return None
    
    def get_orientation(self) -> str:
        """Get predominant orientation for this series."""
        orient_counts = {}
        for img in self.images:
            orient = img.get_orientation()
            orient_counts[orient] = orient_counts.get(orient, 0) + 1
        if orient_counts:
            return max(orient_counts, key=orient_counts.get)
        return 'UNK'
    
    def get_iv_contrast(self) -> str:
        """Get predominant IV contrast status for this series: YES or NO."""
        contrast_counts = {'YES': 0, 'NO': 0}
        for img in self.images:
            contrast_counts[img.get_iv_contrast()] += 1
        return max(contrast_counts, key=contrast_counts.get)
    
    def __repr__(self):
        return f"DICOMSeries(Number: {self.series_number}, Images: {len(self.images)})"


class DICOMDataset:
    """Main container for DICOM data organized by series."""
    
    def __init__(self):
        """Initialize empty DICOM dataset."""
        self.series: Dict[int, DICOMSeries] = {}
    
    def add_image(self, image: DICOMImage):
        """Add an image to the appropriate series."""
        series_num = image.get_series_number()
        
        if series_num not in self.series:
            self.series[series_num] = DICOMSeries(series_num)
        
        self.series[series_num].add_image(image)
    
    def sort_all_series(self):
        """Sort all series by series number and images by instance number."""
        for series in self.series.values():
            series.sort_by_instance_number()
    
    def get_sorted_series_numbers(self) -> List[int]:
        """Get series numbers sorted in ascending order."""
        return sorted(self.series.keys())
    
    def get_series(self, series_number: int) -> Optional[DICOMSeries]:
        """Get a specific series by number."""
        return self.series.get(series_number)
    
    def sort_series_by_tag(self, tag_name: str, reverse: bool = False):
        """Sort images within all series by a specific tag."""
        for series in self.series.values():
            series.sort_by_tag(tag_name, reverse)
    
    def get_all_images(self) -> List[DICOMImage]:
        """Get all images across all series, sorted by series then instance number."""
        all_images = []
        for series_num in self.get_sorted_series_numbers():
            series = self.series[series_num]
            series.sort_by_instance_number()
            all_images.extend(series.images)
        return all_images
    
    def __len__(self):
        return sum(len(series) for series in self.series.values())
    
    def __repr__(self):
        return f"DICOMDataset(Series: {len(self.series)}, Total Images: {len(self)})"


class SeriesPrioritizer:
    """Prioritizes DICOM series based on specific criteria.
    
    Optional user preferences (if not set, that criterion is not used for sorting):
        orient:      Preferred orientation - 'AXL', 'COR', 'SAG', or 'OBL'
        kernel:      Preferred kernel hardness - 'HARD', 'SOFT', or 'MID' (standard)
        iv_contrast: Preferred IV contrast status - 'YES' or 'NO'
        spacing:     Preferred slice spacing in mm (sorts by |actual - target|)
    
    Required flags (if True, series not matching the criterion are excluded entirely):
        orient_required:      If True, only series matching orient are considered
        kernel_required:      If True, only series matching kernel are considered
        iv_contrast_required: If True, only series matching iv_contrast are considered
    
    Spacing range filter (exclude series outside acceptable range):
        min_spacing: Minimum acceptable slice spacing in mm (None = no minimum)
        max_spacing: Maximum acceptable slice spacing in mm (None = no maximum)
    """
    
    def __init__(self, dataset: DICOMDataset, orient=None, kernel=None, iv_contrast=None, spacing=None,
                 orient_required=False, kernel_required=False, iv_contrast_required=False,
                 min_spacing=None, max_spacing=None):
        """Initialize with a DICOM dataset and optional sorting preferences."""
        self.dataset = dataset
        self.orient = orient.upper() if orient else None
        self.kernel = kernel.upper() if kernel else None
        self.iv_contrast = iv_contrast.upper() if iv_contrast else None
        self.spacing = float(spacing) if spacing is not None else None
        self.orient_required = orient_required
        self.kernel_required = kernel_required
        self.iv_contrast_required = iv_contrast_required
        self.min_spacing = float(min_spacing) if min_spacing is not None else None
        self.max_spacing = float(max_spacing) if max_spacing is not None else None
    
    def get_all_series_ranked(self) -> List[DICOMSeries]:
        """Get all series ranked from best to worst based on prioritization criteria."""
        all_series = []
        
        for series_num in self.dataset.get_sorted_series_numbers():
            series = self.dataset.get_series(series_num)
            all_series.append(series)
        
        # --- Hard filter: exclude series that don't meet required criteria ---
        kernel_map = {'HARD': 'Hard', 'SOFT': 'Soft', 'MID': 'Standard'}
        
        if self.orient and self.orient_required:
            before = len(all_series)
            all_series = [s for s in all_series if s.get_orientation() == self.orient]
            excluded = before - len(all_series)
            if excluded:
                print(f"  Filtered out {excluded} series not matching required orient={self.orient}")
        
        if self.kernel and self.kernel_required:
            target_kernel = kernel_map.get(self.kernel, self.kernel)
            before = len(all_series)
            all_series = [s for s in all_series if s.get_kernel_hardness() == target_kernel]
            excluded = before - len(all_series)
            if excluded:
                print(f"  Filtered out {excluded} series not matching required kernel={self.kernel}")
        
        if self.iv_contrast and self.iv_contrast_required:
            before = len(all_series)
            all_series = [s for s in all_series if s.get_iv_contrast() == self.iv_contrast]
            excluded = before - len(all_series)
            if excluded:
                print(f"  Filtered out {excluded} series not matching required iv_contrast={self.iv_contrast}")
        
        # Spacing range filter
        if self.min_spacing is not None or self.max_spacing is not None:
            before = len(all_series)
            def spacing_in_range(s):
                sp = s.get_average_slice_spacing()
                if sp is None:
                    return False  # Unknown spacing excluded when range is specified
                if self.min_spacing is not None and sp < self.min_spacing:
                    return False
                if self.max_spacing is not None and sp > self.max_spacing:
                    return False
                return True
            all_series = [s for s in all_series if spacing_in_range(s)]
            excluded = before - len(all_series)
            if excluded:
                range_str = f"{self.min_spacing or '*'}-{self.max_spacing or '*'}mm"
                print(f"  Filtered out {excluded} series with spacing outside range {range_str}")
        
        if not all_series:
            print("  Warning: No series remaining after applying required filters!")
            return all_series
        
        # Capture preferences for use in sort_key closure
        pref_orient = self.orient
        pref_kernel = self.kernel
        pref_contrast = self.iv_contrast
        pref_spacing = self.spacing
        
        def sort_key(series: DICOMSeries):
            # Priority 0: Exclusion keywords always last (excluded=1, normal=0)
            has_exclusions = 1 if series.has_exclusion_keywords() else 0
            
            # Priority 1: Orientation (only if user specified orient)
            if pref_orient:
                series_orient = series.get_orientation()
                orient_score = 0 if series_orient == pref_orient else 1
            else:
                orient_score = 0  # Not used — all equal
            
            # Priority 2: Kernel (only if user specified kernel)
            if pref_kernel:
                hardness = series.get_kernel_hardness()
                # Map user-facing names to internal names
                kernel_map = {'HARD': 'Hard', 'SOFT': 'Soft', 'MID': 'Standard'}
                target_kernel = kernel_map.get(pref_kernel, pref_kernel)
                kernel_score = 0 if hardness == target_kernel else 1
            else:
                kernel_score = 0  # Not used — all equal
            
            # Priority 3: IV Contrast (only if user specified iv_contrast)
            if pref_contrast:
                series_contrast = series.get_iv_contrast()
                contrast_score = 0 if series_contrast == pref_contrast else 1
            else:
                contrast_score = 0  # Not used — all equal
            
            # Priority 4: Spacing (only if user specified spacing)
            if pref_spacing is not None:
                actual_spacing = series.get_average_slice_spacing()
                if actual_spacing is not None:
                    spacing_score = abs(actual_spacing - pref_spacing)
                else:
                    spacing_score = 999.0  # Unknown spacing = worst
            else:
                spacing_score = 0  # Not used — all equal
            
            return (has_exclusions, orient_score, kernel_score, contrast_score, spacing_score)
        
        all_series.sort(key=sort_key)
        return all_series
    
    def print_all_series_ranked(self):
        """Print all series ranked from best to worst based on prioritization criteria."""
        try:
            all_series = self.get_all_series_ranked()
            
            print(f"\nAll Series Ranked (Best to Worst):")
            print("-" * 90)
            print(f"{'Rank':<4} {'Series':<8} {'Description':<25} {'Axial':<6} {'Z-extent':<10} {'Kernel':<8} {'Spacing':<10} {'Images':<7}")
            print("-" * 90)
            
            for rank, series in enumerate(all_series, 1):
                try:
                    series_desc = series.get_series_description()
                    # Truncate description if too long
                    if len(series_desc) > 24:
                        series_desc = series_desc[:21] + "..."
                    
                    is_axial = "Yes" if series.is_axial() else "No"
                    z_extent = series.get_z_extent()
                    kernel = series.get_kernel_hardness()
                    spacing = series.get_average_slice_spacing()
                    image_count = len(series)
                    
                    z_extent_str = f"{z_extent:.1f}mm" if z_extent else "Unknown"
                    spacing_str = f"{spacing:.1f}mm" if spacing else "Unknown"
                    
                    # Add priority indicators for best matches
                    priority_indicator = ""
                    if series.is_axial():
                        z_extent_val = z_extent if z_extent else 0
                        if 300.0 <= z_extent_val <= 450.0:
                            priority_indicator = " ⭐" if kernel == "Hard" else " ✓"
                    
                    print(f"{rank:<4} {series.series_number:<8} {series_desc:<25} {is_axial:<6} "
                          f"{z_extent_str:<10} {kernel:<8} {spacing_str:<10} {image_count:<7}{priority_indicator}")
                          
                except Exception as e:
                    print(f"Error processing Series {series.series_number}: {e}")
                    
        except Exception as e:
            print(f"Error in series ranking: {e}")
            import traceback
            traceback.print_exc()


def find_dicom_files(directory: str) -> List[Path]:
    """Recursively find all DICOM files in directory and subdirectories."""
    dicom_files = []
    directory_path = Path(directory)
    
    if not directory_path.exists():
        print(f"Directory not found: {directory}")
        return dicom_files
    
    # Common DICOM file extensions
    dicom_extensions = ('.dcm', '.dicom', '.ima', '.img')
    
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = Path(root) / file
            
            # Check by extension first
            if file.lower().endswith(dicom_extensions):
                dicom_files.append(file_path)
            # Also check files without extension (common in DICOM)
            elif '.' not in file:
                # Try to verify it's actually a DICOM file
                if is_dicom_file(file_path):
                    dicom_files.append(file_path)
    
    return dicom_files


def is_dicom_file(file_path: Path) -> bool:
    """Check if a file is a DICOM file by reading its header."""
    if not HAS_PYDICOM:
        return False
        
    try:
        # Try to read just the preamble and check for DICM magic bytes
        with open(file_path, 'rb') as f:
            f.seek(128)  # Skip preamble
            magic = f.read(4)
            return magic == b'DICM'
    except (OSError, IOError):
        return False


def extract_dicom_tags(dicom_path: Path) -> Optional[Dict[str, Any]]:
    """Extract required DICOM tags from a DICOM file."""
    if not HAS_PYDICOM:
        return None
        
    try:
        ds = pydicom.dcmread(dicom_path, stop_before_pixels=True)
        
        # Extract the required tags
        extracted_tags = {}
        
        # Map DICOM tag names to their hex values
        tag_mapping = {
            'SOPClassUID': (0x0008, 0x0016),
            'ImageOrientationPatient': (0x0020, 0x0037),
            'SliceThickness': (0x0018, 0x0050),
            'ReconstructionAlgorithm': (0x0018, 0x9314),  # or 0x0018, 0x1210
            'ContrastBolusAgent': (0x0018, 0x0010),
            'ContrastBolusRoute': (0x0018, 0x1040),
            'ContrastBolusVolume': (0x0018, 0x1041),
            'SeriesDescription': (0x0008, 0x103e),
            'ExamDescription': (0x0008, 0x1030),  # StudyDescription
            'Modality': (0x0008, 0x0060),
            'ConversionType': (0x0008, 0x0064),
            'SeriesNumber': (0x0020, 0x0011),
            'InstanceNumber': (0x0020, 0x0013),
            'BodyPartExamined': (0x0018, 0x0015),
            'SC_Date': (0x0008, 0x0020),  # StudyDate
            'SliceLocation': (0x0020, 0x1041),
            'Manufacturer': (0x0008, 0x0070),
            'ManufacturerModelName': (0x0008, 0x1090),
            'ImagePositionPatient': (0x0020, 0x0032)
        }
        
        # Extract each required tag
        for tag_name, (group, element) in tag_mapping.items():
            extracted_tags[tag_name] = get_ds(ds, group, element)
        
        # Create additional JSON data for improved processing
        json_data = {}
        
        # Add commonly needed tags to JSON data
        additional_tags = {
            'ConvolutionKernel': (0x0018, 0x1210),
            'FilterType': (0x0018, 0x1160),
            'SpacingBetweenSlices': (0x0018, 0x0088),
            'ContrastBolusAgent': (0x0018, 0x0010),
            'ContrastBolusRoute': (0x0018, 0x1040),
            'ContrastBolusVolume': (0x0018, 0x1041),
            'SeriesInstanceUID': (0x0020, 0x000e)
        }
        
        for tag_name, (group, element) in additional_tags.items():
            value = get_ds(ds, group, element)
            if value and value != '-1':
                json_data[tag_name] = value
        
        # Special handling for ImageOrientationPatient - convert to list
        try:
            if hasattr(ds, 'ImageOrientationPatient') and ds.ImageOrientationPatient:
                json_data['ImageOrientationPatient'] = list(ds.ImageOrientationPatient)
                extracted_tags['ImageOrientationPatient'] = str(list(ds.ImageOrientationPatient))
        except:
            pass
            
        # Special handling for ImagePositionPatient - convert to list  
        try:
            if hasattr(ds, 'ImagePositionPatient') and ds.ImagePositionPatient:
                json_data['ImagePositionPatient'] = list(ds.ImagePositionPatient)
                extracted_tags['ImagePositionPatient'] = str(list(ds.ImagePositionPatient))
        except:
            pass
        
        # Add the JSON data as a string (as expected by the CSV format)
        extracted_tags['VD.JSON:TAGS'] = json.dumps(json_data) if json_data else '{}'
        
        return extracted_tags
        
    except Exception as e:
        print(f"Error processing DICOM file {dicom_path}: {e}")
        return None


def create_csv_from_dicom_files(directory: str, csv_output_path: str) -> bool:
    """Scan directory for DICOM files and create a CSV with extracted tags."""
    if not HAS_PYDICOM:
        print("Error: pydicom is required to parse DICOM files. Please install with: pip install pydicom")
        return False
    
    print(f"Scanning for DICOM files in: {directory}")
    dicom_files = find_dicom_files(directory)
    
    if not dicom_files:
        print(f"No DICOM files found in {directory}")
        return False
    
    print(f"Found {len(dicom_files)} DICOM files. Processing...")
    
    # Define CSV columns (required tags + VD.JSON:TAGS)
    csv_columns = DICOMImage.REQUIRED_TAGS + ['VD.JSON:TAGS']
    
    processed_count = 0
    
    try:
        with open(csv_output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
            writer.writeheader()
            
            for dicom_file in dicom_files:
                try:
                    tags = extract_dicom_tags(dicom_file)
                    if tags:
                        # Ensure all required columns are present
                        row = {}
                        for col in csv_columns:
                            row[col] = tags.get(col, '')
                        
                        writer.writerow(row)
                        processed_count += 1
                        
                        if processed_count % 100 == 0:
                            print(f"Processed {processed_count} files...")
                            
                except Exception as e:
                    print(f"Error processing {dicom_file}: {e}")
                    continue
    
        print(f"Successfully created CSV with {processed_count} DICOM entries: {csv_output_path}")
        return True
        
    except Exception as e:
        print(f"Error creating CSV file: {e}")
        return False


def find_csv_files(directory: str) -> List[Path]:
    """Find CSV files in the given directory."""
    directory_path = Path(directory)
    if not directory_path.exists():
        return []
    
    csv_files = list(directory_path.glob('*.csv'))
    return csv_files


def compute_new_info():
    """Placeholder function called after processing CSV file."""
    print("Computing new information...")
    pass


def process_csv_file(csv_file_path: str, dataset: DICOMDataset) -> bool:
    """Process a single CSV file and add its data to the dataset."""
    try:
        with open(csv_file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            images_processed = 0
            
            for row in reader:
                # Create image data dictionary from CSV row
                image_data = {}
                
                # Map CSV columns to DICOM tags
                for tag in DICOMImage.REQUIRED_TAGS:
                    image_data[tag] = row.get(tag, None)
                
                # Parse JSON data from VD.JSON:TAGS column
                json_data = {}
                json_str = row.get('VD.JSON:TAGS', '')
                if json_str:
                    try:
                        json_data = json.loads(json_str)
                    except json.JSONDecodeError:
                        # If JSON parsing fails, continue without JSON data
                        pass
                
                image = DICOMImage(image_data, json_data)
                dataset.add_image(image)
                images_processed += 1
        
        print(f"Processed {images_processed} images from {csv_file_path}")
        return True
        
    except FileNotFoundError:
        print(f"Error: File not found: {csv_file_path}")
        return False
    except csv.Error as e:
        print(f"Error: Invalid CSV in {csv_file_path}: {e}")
        return False
    except Exception as e:
        print(f"Error processing {csv_file_path}: {e}")
        return False


def copy_best_series_dicoms(input_dir: str, series_uid: str, copy_dir: str):
    """Copy all DICOM files matching the given SeriesInstanceUID to copy_dir."""
    if not HAS_PYDICOM:
        print("Error: pydicom is required to copy DICOM files. Please install with: pip install pydicom")
        return 0
    
    copy_dir_path = Path(copy_dir)
    copy_dir_path.mkdir(parents=True, exist_ok=True)
    
    # Clear any existing files in copy_dir
    existing_files = list(copy_dir_path.iterdir())
    if existing_files:
        for f in existing_files:
            if f.is_file():
                f.unlink()
        print(f"Cleared {len([f for f in existing_files if not f.is_dir()])} existing files from {copy_dir}")
    
    dicom_files = find_dicom_files(input_dir)
    if not dicom_files:
        print(f"No DICOM files found in {input_dir} to copy.")
        return 0
    
    copied_count = 0
    print(f"\nScanning {len(dicom_files)} DICOM files for SeriesInstanceUID: {series_uid}")
    
    for dicom_file in dicom_files:
        try:
            ds = pydicom.dcmread(dicom_file, stop_before_pixels=True)
            file_uid = get_ds(ds, 0x0020, 0x000e)
            if file_uid and str(file_uid).strip() == series_uid.strip():
                dest = copy_dir_path / dicom_file.name
                # Handle duplicate filenames by appending a counter
                if dest.exists():
                    stem = dest.stem
                    suffix = dest.suffix
                    counter = 1
                    while dest.exists():
                        dest = copy_dir_path / f"{stem}_{counter}{suffix}"
                        counter += 1
                shutil.copy2(dicom_file, dest)
                copied_count += 1
        except Exception as e:
            # Skip files we can't read
            continue
    
    print(f"Copied {copied_count} DICOM files to {copy_dir}")
    return copied_count


def write_results_excel(results: List[Dict], output_path: str):
    """Write ranked series results to an Excel file.
    
    Args:
        results: List of dicts with keys: Exam, Rank, Series, Description,
                 Orientation, Z_Extent, Kernel, Spacing, Contrast, Images, Selected
        output_path: Path to write the .xlsx file
    """
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not installed. Cannot write Excel file.")
        print("  Install with: pip install openpyxl")
        return
    
    if not results:
        print("No results to write to Excel.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Series Rankings"
    
    # Define columns
    columns = ['Exam', 'Rank', 'Series', 'Description', 'Orientation', 'Z-Extent (mm)',
               'Kernel', 'Spacing (mm)', 'IV Contrast', 'Images', 'Selected']
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Data styling
    center_align = Alignment(horizontal='center')
    selected_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    filtered_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Write data rows
    for row_idx, result in enumerate(results, 2):
        row_data = [
            result.get('Exam', ''),
            result.get('Rank', ''),
            result.get('Series', ''),
            result.get('Description', ''),
            result.get('Orientation', ''),
            result.get('Z_Extent', ''),
            result.get('Kernel', ''),
            result.get('Spacing', ''),
            result.get('Contrast', ''),
            result.get('Images', ''),
            result.get('Selected', ''),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 4:  # Don't center Description
                cell.alignment = center_align
        
        # Highlight selected row green, filtered-out rows red
        if result.get('Selected') == 'YES':
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = selected_fill
        elif result.get('Selected') == 'FILTERED':
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = filtered_fill
    
    # Auto-fit column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(results) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"\nResults written to: {output_path}")


def main(input_path, copy_dir=None, orient=None, kernel=None, iv_contrast=None, desired_spacing=None,
         orient_required=False, kernel_required=False, iv_contrast_required=False,
         min_spacing=None, max_spacing=None):
    """Main processing function.
    
    Args:
        input_path: Path to a directory of DICOM files or a CSV file.
        copy_dir: If set, copy best series DICOM files to this subdirectory.
        orient: Preferred orientation - 'AXL', 'COR', 'SAG', or 'OBL'.
        kernel: Preferred kernel - 'HARD', 'SOFT', or 'MID'.
        iv_contrast: Preferred IV contrast - 'YES' or 'NO'.
        desired_spacing: Preferred slice spacing in mm (float) - sorts by |actual - target|.
        orient_required: If True, exclude series not matching orient.
        kernel_required: If True, exclude series not matching kernel.
        iv_contrast_required: If True, exclude series not matching iv_contrast.
        min_spacing: Minimum acceptable slice spacing in mm (excludes below).
        max_spacing: Maximum acceptable slice spacing in mm (excludes above).
    """

    input_path_obj = Path(input_path)
    
    if not input_path_obj.exists():
        print(f"Error: Path not found: {input_path}")
        sys.exit(1)

    dataset = DICOMDataset()
    csv_file_to_process = None
    
    try:
        if input_path_obj.is_file():
            # Input is a file - assume it's a CSV
            if input_path_obj.suffix.lower() == '.csv':
                csv_file_to_process = input_path
                print(f"Processing CSV file: {csv_file_to_process}")
            else:
                print(f"Error: Input file must be a CSV file, got: {input_path_obj.suffix}")
                sys.exit(1)
                
        elif input_path_obj.is_dir():
            # Input is a directory - check for CSV files first
            print(f"Input is directory: {input_path}")
            csv_files = find_csv_files(input_path)
            
            if csv_files:
                print(f"Found {len(csv_files)} CSV file(s): {[f.name for f in csv_files]}")
                
                if len(csv_files) == 1:
                    csv_file_to_process = str(csv_files[0])
                    print(f"Using CSV file: {csv_file_to_process}")
                else:
                    print("Multiple CSV files found. Please specify which one to use:")
                    for i, csv_file in enumerate(csv_files, 1):
                        print(f"  {i}. {csv_file.name}")
                    
                    try:
                        choice = input("Enter choice (1-{}) or 'all' to process all: ".format(len(csv_files)))
                    except EOFError:
                        print("Non-interactive environment detected. Using first CSV file.")
                        choice = "1"
                    
                    if choice.lower() == 'all':
                        # Process all CSV files
                        for csv_file in csv_files:
                            print(f"\n=== Processing {csv_file.name} ===")
                            temp_dataset = DICOMDataset()
                            if process_csv_file(str(csv_file), temp_dataset):
                                temp_dataset.sort_all_series()
                                compute_new_info()
                                prioritizer = SeriesPrioritizer(temp_dataset, orient=orient, kernel=kernel, iv_contrast=iv_contrast, spacing=desired_spacing, orient_required=orient_required, kernel_required=kernel_required, iv_contrast_required=iv_contrast_required, min_spacing=min_spacing, max_spacing=max_spacing)
                                prioritizer.print_all_series_ranked()
                                if copy_dir  != None:
                                    prioritizer.copy_best_series(copy_dir)
                            else:
                                print(f"Failed to process CSV file: {csv_file}")
                        return []
                    else:
                        try:
                            choice_idx = int(choice) - 1
                            if 0 <= choice_idx < len(csv_files):
                                csv_file_to_process = str(csv_files[choice_idx])
                                print(f"Using CSV file: {csv_file_to_process}")
                            else:
                                print("Invalid choice")
                                sys.exit(1)
                        except ValueError:
                            print("Invalid choice")
                            sys.exit(1)
            else:
                # No CSV files found - scan for DICOM files and create CSV
                print("No CSV files found in directory. Scanning for DICOM files...")
                
                # Create CSV filename based on directory name
                csv_filename = f"{input_path_obj.name}_dicom_tags.csv"
                csv_output_path = input_path_obj / csv_filename
                
                if create_csv_from_dicom_files(input_path, str(csv_output_path)):
                    csv_file_to_process = str(csv_output_path)
                    print(f"Created CSV file: {csv_file_to_process}")
                else:
                    print("Failed to create CSV from DICOM files")
                    sys.exit(1)
        else:
            print(f"Error: Input path is neither a file nor directory: {input_path}")
            sys.exit(1)
        
        # Process the CSV file
        if csv_file_to_process and process_csv_file(csv_file_to_process, dataset):
            dataset.sort_all_series()
            compute_new_info()
            
            print(f"\nProcessing complete!")
            print(f"Final dataset: {dataset}")
            
            # Run series prioritization
            prioritizer = SeriesPrioritizer(dataset, orient=orient, kernel=kernel, iv_contrast=iv_contrast, spacing=desired_spacing, orient_required=orient_required, kernel_required=kernel_required, iv_contrast_required=iv_contrast_required, min_spacing=min_spacing, max_spacing=max_spacing)
            prioritizer.print_all_series_ranked()
            
            # Build results for Excel export
            exam_name = input_path_obj.name
            ranked_series = prioritizer.get_all_series_ranked()
            ranked_nums = {s.series_number for s in ranked_series}
            best_num = ranked_series[0].series_number if ranked_series else None
            results = []
            
            # Add ranked series
            for rank, series in enumerate(ranked_series, 1):
                z_ext = series.get_z_extent()
                sp = series.get_average_slice_spacing()
                results.append({
                    'Exam': exam_name,
                    'Rank': rank,
                    'Series': series.series_number,
                    'Description': series.get_series_description(),
                    'Orientation': series.get_orientation(),
                    'Z_Extent': round(z_ext, 1) if z_ext else 'Unknown',
                    'Kernel': series.get_kernel_hardness(),
                    'Spacing': round(sp, 1) if sp else 'Unknown',
                    'Contrast': series.get_iv_contrast(),
                    'Images': len(series),
                    'Selected': 'YES' if series.series_number == best_num else '',
                })
            
            # Add filtered-out series
            for series_num in dataset.get_sorted_series_numbers():
                if series_num not in ranked_nums:
                    series = dataset.get_series(series_num)
                    z_ext = series.get_z_extent()
                    sp = series.get_average_slice_spacing()
                    results.append({
                        'Exam': exam_name,
                        'Rank': '-',
                        'Series': series.series_number,
                        'Description': series.get_series_description(),
                        'Orientation': series.get_orientation(),
                        'Z_Extent': round(z_ext, 1) if z_ext else 'Unknown',
                        'Kernel': series.get_kernel_hardness(),
                        'Spacing': round(sp, 1) if sp else 'Unknown',
                        'Contrast': series.get_iv_contrast(),
                        'Images': len(series),
                        'Selected': 'FILTERED',
                    })
            
            # Copy best series DICOM files if copy_dir is specified
            if copy_dir:
                if ranked_series:
                    best_series = ranked_series[0]
                    best_uid = best_series.get_series_instance_uid()
                    if best_uid:
                        print(f"\nBest series: #{best_series.series_number} "
                              f"({best_series.get_series_description()})")
                        print(f"SeriesInstanceUID: {best_uid}")
                        
                        if input_path_obj.is_dir():
                            copy_best_series_dicoms(input_path, best_uid, os.path.join(input_path, copy_dir))
                        else:
                            print("Warning: copy_dir requires input_path to be a directory "
                                  "containing DICOM files, not a CSV file.")
                    else:
                        print("Warning: Could not determine SeriesInstanceUID for best series. "
                              "Ensure VD.JSON:TAGS contains SeriesInstanceUID.")
                else:
                    print("No series found to copy.")
            
            return results
        else:
            print(f"Failed to process CSV file: {csv_file_to_process}")
            sys.exit(1)
    
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    return []


if __name__ == "__main__":

    # if len(sys.argv) < 2:
    #     print("Usage: python pick_best_ct_head.py <input_directory_or_csv_file> [copy_dir]")
    #     print("")
    #     print("Head CT Series Prioritization Tool")
    #     print("")
    #     print("If input is a directory:")
    #     print("  - Checks for existing CSV files first")
    #     print("  - If CSV found, processes it (prompts if multiple CSV files)")
    #     print("  - If no CSV found, scans recursively for DICOM files and creates CSV")
    #     print("  - Requires pydicom for DICOM parsing: pip install pydicom")
    #     print("")
    #     print("If input is a CSV file:")
    #     print("  - Processes the CSV file directly")
    #     print("  - CSV must contain DICOM tag columns and VD.JSON:TAGS column")
    #     print("")
    #     print("Optional copy_dir:")
    #     print("  - If provided, copies DICOM files from the best series to this directory")
    #     print("  - Requires input_path to be a directory containing DICOM files")
    #     print("")
    #     print("Output: Prioritized list of head CT series with quality indicators")
    #     sys.exit(1)
    
    # input_path = sys.argv[1]
    # copy_dir = sys.argv[2] if len(sys.argv) > 2 else None

    input_path = '/Users/bje01/Desktop/CQ_Test'
    input_path = '/Volumes/OWC Express 1M2/Images/CQ500_Brain_Hemorrhage_Dataset/CT_DICOMs'

    all_results = []
    subs = os.listdir(input_path)
    for sub in sorted(subs):
        sub_dir = os.path.join(input_path, sub)
        if os.path.isdir(sub_dir):
            print(f"\n{'='*80}")
            print(f"Processing: {sub}")
            print(f"{'='*80}")
            # Sorting Preferences (set to None to disable that criterion):
            #   orient:      Preferred orientation — 'AXL' (axial), 'COR' (coronal), 'SAG' (sagittal), 'OBL' (oblique)
            #   kernel:      Preferred kernel — 'HARD', 'SOFT', or 'MID' (standard)
            #   iv_contrast: Preferred IV contrast — 'YES' (post-contrast) or 'NO' (pre-contrast)
            #   desired_spacing: Preferred slice spacing in mm (float) — sorts by |actual - target|
            #   copy_dir:    Subdirectory name for copying best series DICOM files (relative to exam folder)
            # Required flags (set to True to EXCLUDE series that don't match):
            #   orient_required:      Only consider series matching orient
            #   kernel_required:      Only consider series matching kernel
            #   iv_contrast_required: Only consider series matching iv_contrast
            # Spacing range filter (set to None to disable):
            #   min_spacing: Minimum acceptable slice spacing in mm
            #   max_spacing: Maximum acceptable slice spacing in mm
            results = main(sub_dir, copy_dir='NII',
                 orient='AXL',       orient_required=True,
                 kernel='SOFT',      kernel_required=False,
                 iv_contrast=None,   iv_contrast_required=True,
                 desired_spacing=3,
                 min_spacing=2.0, max_spacing=6.0)
            if results:
                all_results.extend(results)
    
    # Write combined Excel summary
    if all_results:
        excel_path = os.path.join(input_path, 'series_rankings.xlsx')
        write_results_excel(all_results, excel_path)